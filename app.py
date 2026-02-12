import streamlit as st
import google.generativeai as genai
import tempfile
import os
import json
import re
import time
import io
import random
import base64
import pandas as pd
import streamlit.components.v1 as components
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from datetime import datetime, timedelta, timezone

# --- 1. CONFIGURACIÓN E INICIO ---
VER_SISTEMA = "v60.0 (Auto-Scan + Lectura Garantizada)"
ADMIN_USER = "1723623011"
ADMIN_PASS_MASTER = "9994915010022"

st.set_page_config(
    page_title="SIGD DINIC",
    layout="wide",
    page_icon="🛡️",
    initial_sidebar_state="expanded"
)

# --- 2. ESTILOS VISUALES (PREMIUM) ---
st.markdown("""
    <style>
    .main-header { background-color: #0E2F44; padding: 20px; border-radius: 10px; color: white; text-align: center; margin-bottom: 20px; border-bottom: 4px solid #D4AF37; }
    .main-header h1 { margin: 0; font-size: 2.2rem; font-weight: 800; }
    .metric-card { background-color: #f8f9fa !important; border-radius: 10px; padding: 15px; text-align: center; border: 1px solid #dee2e6; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
    .metric-card h3 { color: #0E2F44 !important; font-size: 2rem; margin: 0; font-weight: 800; }
    .metric-card p { color: #555 !important; font-size: 0.9rem; margin: 0; font-weight: 600; }
    .login-container { max-width: 400px; margin: auto; padding: 40px; background-color: #ffffff; border-radius: 15px; box-shadow: 0 10px 25px rgba(0,0,0,0.1); text-align: center; border-top: 5px solid #0E2F44; }
    div.stButton > button { width: 100%; border-radius: 5px; font-weight: bold; }
    </style>
""", unsafe_allow_html=True)

# --- 3. FUNCIONES UTILITARIAS (AL INICIO) ---
def get_hora_ecuador(): 
    return datetime.now(timezone(timedelta(hours=-5)))

def preservar_bordes(cell, fill_obj):
    if cell.border and (cell.border.left.style or cell.border.top.style): new_border = copy(cell.border)
    else:
        thin = Side(border_style="thin", color="000000")
        new_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell.border = new_border; cell.fill = fill_obj
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

def extract_json_safe(text):
    """Extrae JSON incluso si la IA devuelve texto adicional"""
    try: return json.loads(text)
    except:
        match = re.search(r"(\{.*\}|\[.*\])", text, re.DOTALL)
        if match:
            try: 
                data = json.loads(match.group(1))
                if isinstance(data, list) and len(data) > 0: return data[0]
                elif isinstance(data, dict): return data
            except: return {}
        return {}

def limpiar_codigo_prioridad(texto):
    """Regla G7: PN-DIGIN-QX... sin 'Oficio Nro'"""
    if not texto: return ""
    match = re.search(r"(PN-[\w\-\.]+)", str(texto).upper())
    if match: return match.group(1).strip()
    return str(texto).strip()

def extraer_unidad_f7(texto_codigo):
    """Regla F7: Lo que está entre PN- y -QX"""
    if not texto_codigo: return "DINIC"
    match = re.search(r"PN-[\(]?([A-Z0-9]+)[\)]?-QX", str(texto_codigo).upper())
    if match: return match.group(1).strip()
    return "DINIC"

def determinar_sale_no_sale(destinos_str):
    unidades_externas = ["UCAP", "UNDECOF", "UDAR", "DIGIN", "DNATH", "COMANDO GENERAL", "OTRAS DIRECCIONES"]
    destinos_upper = destinos_str.upper()
    for u in unidades_externas:
        if u in destinos_upper: return "SI"
    return "NO"

def get_img_as_base64(file_path):
    try:
        with open(file_path, "rb") as f: return base64.b64encode(f.read()).decode()
    except: return ""

def get_logo_html(width="120px"):
    img_path = "Captura.JPG"
    b64 = get_img_as_base64(img_path)
    if b64: return f'<img src="data:image/jpeg;base64,{b64}" style="width:{width}; margin-bottom:15px;">'
    return f'<img src="https://upload.wikimedia.org/wikipedia/commons/2/25/Escudo_Policia_Nacional_del_Ecuador.png" style="width:{width}; margin-bottom:15px;">'

def frases_curiosas():
    return random.choice(["Analizando...", "Extrayendo datos...", "Verificando códigos...", "Conectando con IA..."])

def generar_html_contrato(datos_usuario, img_b64):
    fecha_hora = get_hora_ecuador().strftime("%Y-%m-%d %H:%M:%S")
    return f"""<div style='font-family:Arial; padding:20px; border:1px solid black;'><h2>ACTA</h2><p>Usuario: {datos_usuario.get('grado')} {datos_usuario.get('nombre')}</p><p>Fecha: {fecha_hora}</p><img src='data:image/png;base64,{img_b64}' width='150'></div>"""

def get_generador_policial_html():
    return """<!DOCTYPE html><html><body><h3>Generador Policial</h3><button>Descargar PDF</button></body></html>"""

# --- 4. GESTIÓN DE MODELOS (AUTO-ESCANEO PARA ARREGLAR 404) ---
def obtener_modelo_disponible():
    """Pregunta a Google qué modelos hay y devuelve el mejor disponible"""
    try:
        listado = genai.list_models()
        modelos_validos = []
        for m in listado:
            if 'generateContent' in m.supported_generation_methods:
                modelos_validos.append(m.name)
        
        # Prioridad: Flash -> Pro 1.5 -> Pro 1.0 -> Cualquiera
        if not modelos_validos: return "models/gemini-1.5-flash" # Fallback por si acaso
        
        # Buscar Flash
        flash = next((m for m in modelos_validos if "flash" in m), None)
        if flash: return flash
        
        # Buscar 1.5 Pro
        pro15 = next((m for m in modelos_validos if "1.5-pro" in m), None)
        if pro15: return pro15
        
        # Devolver el primero que encuentre
        return modelos_validos[0]
    except:
        return "models/gemini-1.5-flash" # Fallback si falla el listado

def invocar_ia_segura(content):
    if 'genai_model' not in st.session_state:
        raise Exception("IA no inicializada")
    
    max_retries = 3
    for i in range(max_retries):
        try:
            return st.session_state.genai_model.generate_content(content)
        except Exception as e:
            if "429" in str(e): # Saturación
                time.sleep(2)
                continue
            # Si es otro error, reintentar una vez más tras pausa
            time.sleep(1)
            
    # Último intento
    return st.session_state.genai_model.generate_content(content)

# --- 5. LOGICA MATRIZ (ESTRICTA) ---
def generar_fila_matriz(tipo, ia_data, manual_data, usuario_turno, paths_files):
    # Extracción
    raw_code_in = ia_data.get("recibido_codigo", "")
    cod_in = limpiar_codigo_prioridad(raw_code_in)
    unidad_f7 = extraer_unidad_f7(raw_code_in)
    
    fecha_in = ia_data.get("recibido_fecha", "")
    remitente_nom = ia_data.get("recibido_remitente_nombre", "")
    remitente_car = ia_data.get("recibido_remitente_cargo", "")
    asunto_in = ia_data.get("recibido_asunto", "")
    resumen_in = ia_data.get("recibido_resumen", "")
    
    dest_out = ia_data.get("respuesta_destinatarios", "")
    raw_code_out = ia_data.get("respuesta_codigo", "")
    cod_out = limpiar_codigo_prioridad(raw_code_out)
    fecha_out = ia_data.get("respuesta_fecha", "")

    # Estados
    estado_s7 = "PENDIENTE"
    has_in = True if (paths_files.get("in") or manual_data.get("G")) else False
    has_out = True if (paths_files.get("out") or manual_data.get("P")) else False
    
    if tipo in ["CONOCIMIENTO", "REASIGNADO", "GENERADO DESDE DESPACHO"]: estado_s7 = "FINALIZADO"
    elif has_in and has_out: estado_s7 = "FINALIZADO"

    str_unidades = manual_data.get("unidades_str", "")
    es_interno = determinar_sale_no_sale(str_unidades)
    if tipo == "CONOCIMIENTO": es_interno = "NO"

    # Estructura Base
    row = {
        "C": fecha_in, "D": remitente_nom, "E": remitente_car, "F": unidad_f7,
        "G": cod_in, "H": fecha_in, "I": asunto_in, "J": resumen_in, "K": usuario_turno,
        "L": "", "M": str_unidades, "N": manual_data.get("tipo_doc_salida", ""),
        "O": "", "P": "", "Q": "", "R": "", "S": estado_s7,
        "T": es_interno, "U": str_unidades, "V": "", "W": "", "X": "", "Y": "", "Z": ""
    }

    if tipo == "TRAMITE NORMAL":
        row["L"] = ""
        row["O"] = dest_out; row["P"] = cod_out; row["Q"] = fecha_out
        row["V"] = cod_out; row["W"] = fecha_out; row["X"] = fecha_out

    elif tipo == "REASIGNADO":
        row["L"] = "REASIGNADO"
        row["O"] = manual_data.get("reasignado_a", "")
        row["P"] = ""; row["V"] = ""
        row["Q"] = fecha_in; row["W"] = fecha_in; row["X"] = fecha_in

    elif tipo == "GENERADO DESDE DESPACHO":
        row["L"] = "GENERADO DESDE DESPACHO"
        row["D"] = ""; row["E"] = ""
        row["C"] = fecha_out; row["H"] = fecha_out; row["Q"] = fecha_out; row["W"] = fecha_out; row["X"] = fecha_out
        row["G"] = cod_out; row["P"] = cod_out; row["V"] = cod_out
        row["F"] = extraer_unidad_f7(cod_out)
        row["O"] = dest_out

    elif tipo == "CONOCIMIENTO":
        row["L"] = "CONOCIMIENTO"
        row["M"] = ""; row["O"] = ""; row["P"] = ""; row["U"] = ""; row["V"] = ""
        row["T"] = "NO"
        row["Q"] = fecha_in; row["W"] = fecha_in; row["X"] = fecha_in

    if row["S"] == "PENDIENTE":
        for k in ["O", "P", "Q", "V", "W", "X"]: row[k] = ""

    return row

# --- 6. CARGA DE DATOS ---
USUARIOS_BASE = {
    "0702870460": {"grado": "SGOS", "nombre": "VILLALTA OCHOA XAVIER BISMARK", "activo": True},
    "1715081731": {"grado": "SGOS", "nombre": "MINDA MINDA FRANCISCO GABRIEL", "activo": True},
    "1723623011": {"grado": "CBOS", "nombre": "CARRILLO NARVAEZ JOHN STALIN", "activo": True}
}
UNIDADES_DEFAULT = ["DINIC", "SOPORTE OPERATIVO", "APOYO OPERATIVO", "PLANIFICACION", "JURIDICO", "COMUNICACION", "UCAP", "UNDECOF", "UDAR", "DIGIN", "DNATH"]

DB_FILE = "usuarios_db.json"; CONFIG_FILE = "config_sistema.json"; CONTRATOS_FILE = "contratos_legal.json"; LOGS_FILE = "historial_acciones.json"; LISTAS_FILE = "listas_db.json"

def cargar_json(filepath, default):
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r') as f: return json.load(f)
        except: return default
    return default
def guardar_json(filepath, data):
    try:
        with open(filepath, 'w') as f: json.dump(data, f)
    except: pass

db_usuarios = cargar_json(DB_FILE, USUARIOS_BASE)
for k,v in USUARIOS_BASE.items(): 
    if k not in db_usuarios: db_usuarios[k] = v
db_contratos = cargar_json(CONTRATOS_FILE, {})
db_logs = cargar_json(LOGS_FILE, [])
if not isinstance(db_logs, list): db_logs = []
db_listas = cargar_json(LISTAS_FILE, {"unidades": UNIDADES_DEFAULT, "reasignados": []})
config_sistema = cargar_json(CONFIG_FILE, {"pass_universal": "DINIC2026", "pass_th": "THDINIC123", "base_historica": 1258, "consultas_ia_global": 0})

def registrar_accion(usuario, accion, detalle=""):
    ahora = get_hora_ecuador().strftime("%Y-%m-%d %H:%M:%S")
    db_logs.insert(0, {"fecha": ahora, "usuario": usuario, "accion": accion, "detalle": detalle})
    guardar_json(LOGS_FILE, db_logs)

def guardar_nueva_entrada_lista(tipo, valor):
    if valor and valor not in db_listas[tipo]:
        db_listas[tipo].append(valor); guardar_json(LISTAS_FILE, db_listas)
        if tipo == "unidades": st.session_state.lista_unidades = db_listas["unidades"]
        if tipo == "reasignados": st.session_state.lista_reasignados = db_listas["reasignados"]

def actualizar_presencia(cedula_usuario):
    if cedula_usuario in db_usuarios:
        db_usuarios[cedula_usuario]['ultima_actividad'] = get_hora_ecuador().strftime("%Y-%m-%d %H:%M:%S")
        guardar_json(DB_FILE, db_usuarios)

def get_estado_usuario(cedula):
    user_data = db_usuarios.get(cedula, {})
    last_seen_str = user_data.get('ultima_actividad')
    if not last_seen_str: return "🔴 DESCONECTADO"
    try:
        last_seen = datetime.strptime(last_seen_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone(timedelta(hours=-5)))
        diferencia = (get_hora_ecuador() - last_seen).total_seconds() / 60
        if diferencia < 5: return "🟢 EN LÍNEA"
        else: return "🔴 DESCONECTADO"
    except: return "🔴 ERROR"

# --- 7. VARIABLES DE SESIÓN ---
if 'logged_in' not in st.session_state: st.session_state.logged_in = False
if 'user_role' not in st.session_state: st.session_state.user_role = "" 
if 'usuario_turno' not in st.session_state: st.session_state.usuario_turno = "" 
if 'user_id' not in st.session_state: st.session_state.user_id = ""
if 'registros' not in st.session_state: st.session_state.registros = [] 
if 'edit_index' not in st.session_state: st.session_state.edit_index = None 
if 'docs_procesados_hoy' not in st.session_state: st.session_state.docs_procesados_hoy = 0
if 'consultas_ia' not in st.session_state: st.session_state.consultas_ia = 0
if 'active_module' not in st.session_state: st.session_state.active_module = 'secretario'
if 'lista_unidades' not in st.session_state: st.session_state.lista_unidades = db_listas.get("unidades", UNIDADES_DEFAULT)
if 'lista_reasignados' not in st.session_state: st.session_state.lista_reasignados = db_listas.get("reasignados", [])
if 'active_model_name' not in st.session_state: st.session_state.active_model_name = "Detectando..."

# Persistencia F5
token = st.query_params.get("token", None)
if token and not st.session_state.logged_in and token in db_usuarios:
    st.session_state.logged_in = True; st.session_state.user_id = token; st.session_state.user_role = "admin" if token == ADMIN_USER else "user"
    st.session_state.usuario_turno = f"{db_usuarios[token]['grado']} {db_usuarios[token]['nombre']}"

# --- AUTO-SCAN DE IA (CORAZÓN DE LA SOLUCIÓN) ---
sistema_activo = False
try:
    api_key = st.secrets.get("GEMINI_API_KEY")
    if api_key:
        genai.configure(api_key=api_key)
        # Escanear modelos disponibles y elegir el mejor automáticamente
        if 'genai_model' not in st.session_state:
            modelo_elegido = obtener_modelo_disponible()
            st.session_state.genai_model = genai.GenerativeModel(modelo_elegido)
            st.session_state.active_model_name = modelo_elegido
        sistema_activo = True
except: sistema_activo = False

# ==============================================================================
#  INTERFAZ GRÁFICA PRINCIPAL
# ==============================================================================

if not st.session_state.logged_in:
    c1, c2, c3 = st.columns([1,2,1])
    with c2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(f"""<div class="login-container">{get_logo_html()}<h2 style='color:#0E2F44; margin-bottom: 5px;'>ACCESO SIGD DINIC</h2><p style='color: gray; margin-top: 0;'>Sistema Oficial de Gestión Documental</p></div>""", unsafe_allow_html=True)
        with st.form("login_form"):
            usuario_input = st.text_input("Usuario (Cédula):").strip()
            pass_input = st.text_input("Contraseña:", type="password").strip()
            if st.form_submit_button("INGRESAR AL SISTEMA", type="primary"):
                if usuario_input == ADMIN_USER and pass_input == ADMIN_PASS_MASTER:
                    st.session_state.logged_in = True; st.session_state.user_role = "admin"; st.session_state.user_id = usuario_input
                    admin_data = db_usuarios.get(ADMIN_USER, {"grado": "CBOS.", "nombre": "CARRILLO NARVAEZ JOHN STALIN"})
                    st.session_state.usuario_turno = f"{admin_data['grado']} {admin_data['nombre']}"
                    registrar_accion(st.session_state.usuario_turno, "INICIO SESIÓN ADMIN")
                    actualizar_presencia(usuario_input); st.query_params["token"] = usuario_input; st.rerun()
                elif usuario_input in db_usuarios:
                    user_data = db_usuarios[usuario_input]
                    if pass_input == config_sistema["pass_universal"]:
                        if user_data["activo"]:
                            st.session_state.logged_in = True; st.session_state.user_role = "user"; st.session_state.user_id = usuario_input
                            st.session_state.usuario_turno = f"{user_data['grado']} {user_data['nombre']}"
                            registrar_accion(st.session_state.usuario_turno, "INICIO SESIÓN USUARIO")
                            actualizar_presencia(usuario_input); st.query_params["token"] = usuario_input; st.rerun()
                        else: st.error("🚫 Usuario inactivo.")
                    else: st.error("🚫 Contraseña incorrecta.")
                else: st.error("🚫 Usuario no autorizado.")

else:
    actualizar_presencia(st.session_state.user_id)
    with st.sidebar:
        if os.path.exists("Captura.JPG"): st.image("Captura.JPG", use_container_width=True)
        else: st.image("https://upload.wikimedia.org/wikipedia/commons/2/25/Escudo_Policia_Nacional_del_Ecuador.png", width=100)
        st.markdown("### 👮‍♂️ CONTROL DE MANDO")
        if st.session_state.user_role == "admin": st.markdown("""<div class="admin-badge">🛡️ MODO ADMINISTRADOR<br><span style="font-size: 0.8em; font-weight: normal;">CONTROL TOTAL</span></div>""", unsafe_allow_html=True)
        st.info(f"👤 **{st.session_state.usuario_turno}**")
        st.caption(f"🤖 IA: {st.session_state.active_model_name}") 
        fecha_turno = st.date_input("Fecha Operación:", value=get_hora_ecuador().date())
        st.markdown("---"); st.markdown("### 📂 MÓDULOS")
        if st.button("📝 SECRETARIO/A", use_container_width=True, type="primary" if st.session_state.active_module == 'secretario' else "secondary"): st.session_state.active_module = 'secretario'; st.rerun()
        if st.button("🧠 ASESOR INTELIGENTE", use_container_width=True, type="primary" if st.session_state.active_module == 'asesor' else "secondary"): st.session_state.active_module = 'asesor'; st.rerun()
        if st.button("👤 TALENTO HUMANO", use_container_width=True, type="primary" if st.session_state.active_module == 'th' else "secondary"): st.session_state.active_module = 'th'; st.rerun()
        if st.button("🛡️ ADMINISTRADOR", use_container_width=True, type="primary" if st.session_state.active_module == 'admin' else "secondary"): st.session_state.active_module = 'admin'; st.rerun()
        st.markdown("---")
        if st.button("🔒 CERRAR SESIÓN"): 
            st.session_state.logged_in = False; st.query_params.clear(); st.rerun()

    # --- MÓDULO SECRETARIO ---
    if st.session_state.active_module == 'secretario':
        st.markdown(f'''<div class="main-header"><h1>SIGD DINIC</h1><h3>Módulo Secretario/a - Gestión Documental</h3></div>''', unsafe_allow_html=True)
        
        base_h = config_sistema.get("base_historica", 1258)
        total_d = base_h + len(st.session_state.registros)
        total_ia = config_sistema.get("consultas_ia_global", 0) + st.session_state.consultas_ia
        c1, c2, c3 = st.columns(3)
        c1.markdown(f"<div class='metric-card'><h3>📥 {st.session_state.docs_procesados_hoy}</h3><p>Docs Turno Actual</p></div>", unsafe_allow_html=True)
        c2.markdown(f"<div class='metric-card'><h3>📈 {total_d}</h3><p>Total Histórico</p></div>", unsafe_allow_html=True)
        c3.markdown(f"<div class='metric-card'><h3>🧠 {total_ia}</h3><p>Consultas IA (Global)</p></div>", unsafe_allow_html=True)
        
        with st.expander("⚙️ CONFIGURACIÓN Y RESPALDO RÁPIDO"):
            c_conf1, c_conf2 = st.columns(2)
            with c_conf1:
                safe_user = re.sub(r'[^a-zA-Z0-9]', '_', st.session_state.usuario_turno)
                date_str = get_hora_ecuador().strftime("%Y-%m-%d")
                bk_name = f"RESPALDO_TURNO_{safe_user}_{date_str}.json"
                if st.session_state.registros: 
                    st.download_button("⬇️ RESPALDAR TURNO", json.dumps(st.session_state.registros, default=str), bk_name, "application/json")
                up_bk = st.file_uploader("⬆️ RESTAURAR TURNO", type=['json'])
                if up_bk: 
                    try: 
                        st.session_state.registros = json.load(up_bk); st.session_state.docs_procesados_hoy = len(st.session_state.registros); st.success("¡Restaurado!"); time.sleep(1); st.rerun()
                    except: st.error("Error archivo.")
            with c_conf2:
                if os.path.exists("matriz_maestra.xlsx"):
                    st.success("✅ Matriz Cargada")
                    if st.button("🔄 Cambiar Matriz"): os.remove("matriz_maestra.xlsx"); st.rerun()
                else:
                    up_m = st.file_uploader("Cargar Matriz .xlsx", type=['xlsx'])
                    if up_m: 
                        with open("matriz_maestra.xlsx", "wb") as f: f.write(up_m.getbuffer())
                        st.rerun()
        st.write("")

        if sistema_activo:
            is_edit = st.session_state.edit_index is not None
            idx_edit = st.session_state.edit_index
            reg_edit = st.session_state.registros[idx_edit] if is_edit else None
            
            if is_edit: st.warning(f"✏️ EDITANDO REGISTRO #{idx_edit + 1}"); 
            else: st.info("🆕 NUEVO REGISTRO")
            
            col1, col2 = st.columns([1, 2])
            with col1:
                v_tipo = reg_edit['L'] if (is_edit and reg_edit['L']) else "TRAMITE NORMAL"
                if not v_tipo: v_tipo = "TRAMITE NORMAL"
                tipo_proc = st.selectbox("Tipo Gestión:", ["TRAMITE NORMAL", "REASIGNADO", "GENERADO DESDE DESPACHO", "CONOCIMIENTO"], index=["TRAMITE NORMAL", "REASIGNADO", "GENERADO DESDE DESPACHO", "CONOCIMIENTO"].index(v_tipo))
                v_sal = reg_edit['N'] if (is_edit and reg_edit['N']) else "QUIPUX ELECTRONICO"
                tipo_doc = st.selectbox("Formato Salida:", ["QUIPUX ELECTRONICO", "DOCPOL ELECTRONICO", "FISICO", "DIGITAL", "OTRO"], index=["QUIPUX ELECTRONICO", "DOCPOL ELECTRONICO", "FISICO", "DIGITAL", "OTRO"].index(v_sal) if v_sal else 0)
                
                st.markdown("---"); st.caption("🏢 DEPENDENCIA/as DE DESTINO")
                opts_u = sorted(st.session_state.lista_unidades)
                def_u = [u for u in (reg_edit['M'].split(", ") if is_edit and reg_edit['M'] else []) if u in opts_u]
                u_sel = st.multiselect("Seleccione:", opts_u, default=def_u)
                chk_no = st.checkbox("NINGUNA")
                chk_ot = st.checkbox("✍️ OTRA")
                in_ot = st.text_input("Nueva:") if chk_ot else ""
                list_u = u_sel.copy()
                if in_ot: list_u.append(in_ot.upper())
                str_u = ", ".join(list_u) if not chk_no else ""
                
                dest_reasig = ""
                if tipo_proc == "REASIGNADO":
                    st.markdown("---"); st.markdown("👤 **DESTINATARIO REASIGNADO**")
                    opts_r = ["SELECCIONAR..."] + sorted(st.session_state.lista_reasignados) + ["✍️ NUEVO"]
                    idx_r = opts_r.index(reg_edit["O"]) if (is_edit and reg_edit.get("O") in opts_r) else 0
                    sel_r = st.selectbox("Historial:", opts_r, index=idx_r)
                    in_r_man = st.text_input("Grado y Nombre:", value=reg_edit.get("O","") if is_edit else "")
                    if sel_r == "✍️ NUEVO" or in_r_man: dest_reasig = in_r_man.upper()
                    elif sel_r != "SELECCIONAR...": dest_reasig = sel_r

            with col2:
                d_in = None; d_out = None
                if tipo_proc == "TRAMITE NORMAL":
                    c_in, c_out = st.columns(2)
                    d_in = c_in.file_uploader("1. Doc RECIBIDO (PDF)", ['pdf'])
                    d_out = c_out.file_uploader("2. Doc RESPUESTA (PDF)", ['pdf'])
                elif tipo_proc in ["REASIGNADO", "CONOCIMIENTO"]:
                    d_in = st.file_uploader("1. Doc RECIBIDO (PDF)", ['pdf'])
                elif tipo_proc == "GENERADO DESDE DESPACHO":
                    d_out = st.file_uploader("2. Doc GENERADO (PDF)", ['pdf'])

            if st.button("🔄 ACTUALIZAR" if is_edit else "➕ AGREGAR", type="primary"):
                if not os.path.exists("matriz_maestra.xlsx"): st.error("❌ Falta Matriz.")
                else:
                    process = False
                    if tipo_proc == "TRAMITE NORMAL": process = True if (is_edit or d_in or d_out) else False
                    elif d_in or d_out: process = True
                    
                    if process:
                        with st.spinner(f"⏳ PROCESANDO... {frases_curiosas()}"):
                            try:
                                paths = {"in":None, "out":None}
                                if d_in:
                                    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as t: t.write(d_in.getvalue()); paths["in"] = t.name
                                if d_out:
                                    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as t: t.write(d_out.getvalue()); paths["out"] = t.name
                                
                                files_ia = []
                                if paths["in"]: files_ia.append(genai.upload_file(paths["in"], display_name="In"))
                                if paths["out"]: files_ia.append(genai.upload_file(paths["out"], display_name="Out"))
                                
                                prompt = """
                                Actúa como experto en gestión documental policial (DINIC). Analiza y extrae JSON ESTRICTO.
                                
                                1. CÓDIGO (CRÍTICO): Busca en la esquina superior DERECHA (encabezado). Formato "Oficio Nro. PN-..." o "Memorando...". Extrae TODO el código tal cual.
                                
                                2. DESTINATARIOS (Campo O7):
                                   - UBICACIÓN CLAVE: Busca la sección "PARA:" en la parte SUPERIOR del documento.
                                   - INSTRUCCIÓN: Extrae NOMBRES y GRADOS de esa sección.
                                   - REGLA DE ORO: SI LA LINEA TIENE UN CARGO (Ej: "Jefe de...", "Director...", "Comandante..."), ¡IGNORA ESA LINEA!. SOLO QUIERO EL NOMBRE.
                                   - ¡PROHIBIDO!: NO mires la parte inferior (firma/atentamente). NO extraigas Cargos.
                                
                                3. REMITENTE (Campo D7):
                                   - Busca "DE:" en la cabecera O la firma al final. Extrae GRADO y NOMBRE (Ej: Sgos. Juan Perez).
                                   - IMPORTANTE: SIEMPRE incluye el GRADO.

                                JSON:
                                {
                                    "recibido_fecha": "DD/MM/AAAA",
                                    "recibido_remitente_nombre": "Texto (Grado y Nombre)",
                                    "recibido_remitente_cargo": "Texto",
                                    "recibido_codigo": "Texto",
                                    "recibido_asunto": "Texto",
                                    "recibido_resumen": "Texto",
                                    "respuesta_destinatarios": "Texto (Solo Nombres/Grados de la sección PARA, separados por coma, SIN CARGOS)",
                                    "respuesta_codigo": "Texto",
                                    "respuesta_fecha": "DD/MM/AAAA"
                                }
                                """
                                data_ia = {}
                                if files_ia:
                                    res = invocar_ia_segura([prompt, *files_ia])
                                    txt_clean = res.text.replace("```json", "").replace("```", "")
                                    data_ia = extract_json_safe(txt_clean)
                                    
                                    if not data_ia: st.error("⚠️ La IA no devolvió datos válidos. Revisa el PDF.")
                                
                                final_d = reg_edit.copy() if is_edit else {}
                                man_data = {"unidades_str": str_u, "tipo_doc_salida": tipo_doc, "reasignado_a": dest_reasig, "G": final_d.get("G",""), "P": final_d.get("P","")}
                                row = generar_fila_matriz(tipo_proc, data_ia, man_data, st.session_state.usuario_turno, paths)
                                
                                if in_ot: guardar_nueva_entrada_lista("unidades", in_ot)
                                if dest_reasig: guardar_nueva_entrada_lista("reasignados", dest_reasig)
                                
                                if is_edit: st.session_state.registros[idx_edit] = row; st.session_state.edit_index = None; st.success("✅ Actualizado"); registrar_accion(st.session_state.usuario_turno, f"EDITÓ {row['G']}")
                                else: st.session_state.registros.append(row); st.session_state.docs_procesados_hoy += 1; st.success("✅ Agregado"); registrar_accion(st.session_state.usuario_turno, f"NUEVO {row['G']}")
                                
                                if paths["in"]: os.remove(paths["in"])
                                if paths["out"]: os.remove(paths["out"])
                                st.rerun()
                            except Exception as e: st.error(f"Error Técnico: {e}")
                    else: st.warning("⚠️ Sube documento.")

            if st.session_state.registros:
                st.markdown("#### 📋 Cola de Trabajo")
                if len(st.session_state.registros)>0:
                    inds = [f"#{i+1} | {r['G']} | {r['L']}" for i,r in enumerate(st.session_state.registros)]
                    s_idx = st.selectbox("Ver:", range(len(st.session_state.registros)), format_func=lambda x: inds[x], index=len(st.session_state.registros)-1)
                    st.dataframe(pd.DataFrame([st.session_state.registros[s_idx]]), hide_index=True)
                
                for i, r in enumerate(st.session_state.registros):
                    bg = "#e8f5e9" if r["S"]=="FINALIZADO" else "#ffebee"
                    with st.container():
                        st.markdown(f"<div style='background:{bg}; padding:10px; border-left:5px solid {'green' if r['S']=='FINALIZADO' else 'red'}; margin-bottom:5px; border-radius:5px;'><b>#{i+1}</b> | <b>{r['G']}</b> | {r['L']}</div>", unsafe_allow_html=True)
                        c_e, c_d = st.columns([1,1])
                        if c_e.button("✏️", key=f"e{i}"): st.session_state.edit_index = i; st.rerun()
                        if c_d.button("🗑️", key=f"d{i}"): st.session_state.registros.pop(i); st.session_state.docs_procesados_hoy = max(0, st.session_state.docs_procesados_hoy-1); st.rerun()
                
                if os.path.exists("matriz_maestra.xlsx"):
                    try:
                        wb = load_workbook("matriz_maestra.xlsx"); ws = wb[next((s for s in wb.sheetnames if "CONTROL" in s.upper()), wb.sheetnames[0])]
                        start_row = 7
                        while ws.cell(row=start_row, column=1).value is not None: start_row += 1
                        gf = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                        rf = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        for i, r in enumerate(st.session_state.registros):
                            rw = start_row + i
                            def w(c,v): ws.cell(row=rw, column=c).value = v
                            w(1, i+1); w(2, ""); w(3, r["C"]); w(4, r["D"]); w(5, r["E"]); w(6, r["F"]); w(7, r["G"]); w(8, r["H"]); w(9, r["I"]); w(10, r["J"]); w(11, r["K"]); w(12, r["L"]); w(13, r["M"]); w(14, r["N"]); w(15, r["O"]); w(16, r["P"]); w(17, r["Q"]); w(18, "")
                            cs = ws.cell(row=rw, column=19); cs.value = r["S"]
                            if r["S"]=="FINALIZADO": preservar_bordes(cs, gf)
                            else: preservar_bordes(cs, rf)
                            w(20, r["T"]); w(21, r["U"]); w(22, r["V"]); w(23, r["W"]); w(24, r["X"]); w(25, ""); w(26, "")
                            for c_idx in range(1, 27): 
                                cell = ws.cell(row=rw, column=c_idx)
                                if c_idx != 19: preservar_bordes(cell, PatternFill(fill_type=None))
                        out = io.BytesIO(); wb.save(out); out.seek(0)
                        fn = f"TURNO {fecha_turno.strftime('%d-%m-%y')} {st.session_state.usuario_turno.upper()}.xlsx"
                        st.download_button("📥 DESCARGAR MATRIZ FINAL", out, fn, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
                    except Exception as e: st.error(f"Error Excel: {e}")

    # --- MÓDULO ASESOR ---
    elif st.session_state.active_module == 'asesor':
        st.markdown("### 🧠 Asesor Inteligente")
        if st.session_state.user_id not in db_contratos:
            st.warning("⚠️ Acepte los términos.")
            with st.expander("📜 TÉRMINOS Y CONDICIONES"):
                if st.button("✅ ACEPTAR Y FIRMAR"):
                    db_contratos[st.session_state.user_id] = {"fecha": get_hora_ecuador().strftime("%Y-%m-%d %H:%M:%S"), "foto": "", "usuario": st.session_state.usuario_turno}
                    guardar_json(CONTRATOS_FILE, db_contratos); st.rerun()
        else:
            st.markdown("""<div class="legal-warning">⚠️ AVISO LEGAL: Uso referencial.</div>""", unsafe_allow_html=True)
            if st.session_state.user_id in db_contratos:
                st.download_button("📜 Descargar Contrato", generar_html_contrato(db_usuarios.get(st.session_state.user_id,{}),""), "Contrato.html", "text/html")
            up_as = st.file_uploader("Sube PDF", ['pdf'])
            if up_as and st.button("ANALIZAR"):
                with st.spinner("Analizando..."):
                    try:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as t: t.write(up_as.getvalue()); p = t.name
                        f = genai.upload_file(p, display_name="Consulta")
                        res = invocar_ia_segura(["Actúa como JEFE. Dame Diagnóstico, Criterio y Extracto.", f])
                        st.markdown(res.text)
                        st.session_state.consultas_ia += 1; incrementar_contador_ia()
                        os.remove(p)
                    except Exception as e: st.error(str(e))

    # --- MÓDULO TH ---
    elif st.session_state.active_module == 'th':
        if not st.session_state.th_unlocked:
            st.markdown("### 👤 Talento Humano"); pwd = st.text_input("Contraseña:", type="password")
            if st.button("Ingresar"): 
                if pwd == config_sistema.get("pass_th", "THDINIC123"): st.session_state.th_unlocked = True; st.rerun()
                else: st.error("Incorrecto")
        else:
            components.html(get_generador_policial_html(), height=800, scrolling=True)
            if st.button("Cerrar"): st.session_state.th_unlocked = False; st.rerun()

    # --- MÓDULO ADMINISTRADOR (RESTAURADO) ---
    elif st.session_state.active_module == 'admin':
        st.markdown("### 🛡️ ADMINISTRADOR"); pwd = st.text_input("Contraseña Maestra:", type="password")
        if st.session_state.user_role == "admin" and pwd == ADMIN_PASS_MASTER:
            t1, t2, t3, t4 = st.tabs(["Monitor", "Contratos", "Historial", "Config"])
            with t1:
                data = [{"Usuario": f"{v['grado']} {v['nombre']}", "Estado": get_estado_usuario(k)} for k,v in db_usuarios.items()]
                st.dataframe(pd.DataFrame(data), use_container_width=True)
            with t2:
                if db_contratos:
                    for k, v in db_contratos.items():
                        c1, c2, c3 = st.columns([2,1,1])
                        c1.write(f"{v['usuario']} ({v['fecha']})")
                        c2.download_button("⬇️", generar_html_contrato(db_usuarios.get(k,{}), v["foto"]), f"C_{k}.html", key=f"dl_{k}")
                        if c3.button("🗑️", key=f"del_{k}"): del db_contratos[k]; guardar_json(CONTRATOS_FILE, db_contratos); st.rerun()
                else: st.info("Vacío")
            with t3: st.dataframe(pd.DataFrame(db_logs), use_container_width=True)
            with t4:
                st.caption("Contadores"); val = st.number_input("Consultas IA:", value=config_sistema.get("consultas_ia_global",0))
                if st.button("Guardar IA"): config_sistema["consultas_ia_global"]=val; guardar_json(CONFIG_FILE, config_sistema); st.success("OK")
